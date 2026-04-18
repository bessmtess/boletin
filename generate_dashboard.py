#!/usr/bin/env python3
"""Generate ultra-improved BESS2025 dashboard - MTESS style."""
import argparse
import json
import os
import re
from pathlib import Path
from typing import Optional

import openpyxl

YEARS = ["2020", "2021", "2022", "2023", "2024"]
AGE_GROUP_RE = re.compile(r"^\s*\d+\s*a\s*\d+\s*$", re.IGNORECASE)
AGE_OPEN_RE = re.compile(r"^\s*\d+\s*(?:y\s*m[aá]s|\+)\s*$", re.IGNORECASE)


def years_map(row: list, start_col: int) -> dict:
    out = {}
    for i, year in enumerate(YEARS):
        idx = start_col + i
        out[year] = row[idx] if idx < len(row) and row[idx] is not None else 0
    return out


def is_age_group(label: str) -> bool:
    return bool(AGE_GROUP_RE.match(label) or AGE_OPEN_RE.match(label))


def normalize_age_group(label: str) -> str:
    return re.sub(r"\s+", " ", label.strip())


def resolve_excel_path(cli_path: Optional[str]) -> Path:
    if cli_path:
        return Path(cli_path).expanduser()

    env_path = os.getenv("BESS_EXCEL_PATH")
    if env_path:
        return Path(env_path).expanduser()

    candidates = [
        # Common location in this workspace (repo lives in BESS2025/GITHUB)
        Path(__file__).resolve().parent.parent / "Anexo_Estadístico_Boletin_Seguridad_Social_2025.xlsx",
        # Same folder as this script (if someone copies the Excel into the repo)
        Path(__file__).resolve().parent / "Anexo_Estadístico_Boletin_Seguridad_Social_2025.xlsx",
        # Historical absolute path (kept as last-resort fallback)
        Path(
            "/Users/diegobernardomezabogado/Library/CloudStorage/GoogleDrive-dmeza.py@gmail.com/Mi unidad/BESS_MTESS/Anexo_Estadístico_Boletin_Seguridad_Social_2025.xlsx"
        ),
    ]
    for c in candidates:
        if c.exists():
            return c
    return candidates[0]


parser = argparse.ArgumentParser(description="Generate MTESS-style BESS dashboard (index.html).")
parser.add_argument(
    "--excel-path",
    dest="excel_path",
    default=None,
    help="Ruta al Excel del Anexo Estadístico (xlsx). Si se omite, usa BESS_EXCEL_PATH o rutas comunes.",
)
parser.add_argument(
    "--out",
    dest="out_path",
    default=None,
    help="Ruta de salida del HTML. Por defecto: index.html en este directorio.",
)
args = parser.parse_args()

excel_path = resolve_excel_path(args.excel_path)
if not excel_path.exists():
    raise SystemExit(f"Excel no encontrado: {excel_path} (use --excel-path o BESS_EXCEL_PATH)")

wb = openpyxl.load_workbook(excel_path, data_only=True)

# Nota: En el Excel actual, las secciones de CAJUBI y Caja Municipal están numeradas al revés
# respecto al mapeo original. Por eso se invierte aquí (corrige la app web).
cajas_map = {
    "1": "IPS",
    "2": "Caja Fiscal",
    "3": "Caja ANDE",
    "4": "Caja Bancaria",
    "5": "CAJUBI",
    "6": "Caja Ferroviaria",
    "7": "Caja Municipal",
    "8": "Caja Parlamentaria",
}

all_data = {}
for name in wb.sheetnames:
    if name in ['Índice','FichaTécnica']: continue
    ws = wb[name]
    all_data[name] = [[v for v in r] for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)]

summary = {}
for sec_num, caja_name in cajas_map.items():
    cd = {}
    sk = f'{sec_num}.1'
    if sk in all_data:
        for row in all_data[sk]:
            label = str(row[1]).strip().lower() if len(row) > 1 and row[1] is not None else ""
            if label in {"activos", "cotizantes", "cotizantes activos"}:
                cd["activos"] = years_map(row, 2)
            elif label in {"pasivos", "jubilados/pensionados", "jubilados", "pensionados"}:
                cd["pasivos"] = years_map(row, 2)
            elif label and ("relación" in label or "relacion" in label):
                cd["relacion"] = years_map(row, 2)
    sk = f'{sec_num}.2'
    if sk in all_data:
        tipos = {}
        for row in all_data[sk]:
            if row[1] in ['Vejez','Pensión/Sobrevivencia','Invalidez','Otros']:
                tipos[row[1]]={'2020':row[2],'2021':row[3],'2022':row[4],'2023':row[5],'2024':row[6]}
            elif row[1] and 'Total' in str(row[1]):
                tipos['Total']={'2020':row[2],'2021':row[3],'2022':row[4],'2023':row[5],'2024':row[6]}
        cd['jubilados_tipo']=tipos
    sk = f'{sec_num}.3'
    if sk in all_data:
        sx={'activos':{},'pasivos':{}}; cg=None
        for row in all_data[sk]:
            block = str(row[1]).strip().lower() if len(row) > 1 and row[1] is not None else ""
            if block in {"activos", "cotizantes", "cotizantes activos"}:
                cg = "activos"
            elif block in {"pasivos", "jubilados/pensionados", "jubilados", "pensionados"}:
                cg = "pasivos"

            sex = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            if cg and sex == "Hombres":
                sx[cg]["Hombres"] = years_map(row, 3)
            elif cg and sex == "Mujeres":
                sx[cg]["Mujeres"] = years_map(row, 3)
            elif cg and sex == "Total":
                sx[cg]["Total"] = years_map(row, 3)
        cd['por_sexo']=sx
    sk = f'{sec_num}.4'
    if sk in all_data:
        sal={}
        for row in all_data[sk]:
            if row[1] and str(row[1]).strip() in ['2020','2021','2022','2023','2024']:
                sal[str(row[1]).strip()]={'salario_activo':row[2],'jubilacion_promedio':row[3],'ratio':row[4]}
        cd['salarios']=sal
    for suffix, key in [("5", "activos_edad"), ("7", "vejez_edad"), ("6", "salario_edad")]:
        sk = f'{sec_num}.{suffix}'
        if sk in all_data:
            items = []
            seen_groups: set[str] = set()
            for row in all_data[sk]:
                label_raw = str(row[1]) if len(row) > 1 and row[1] is not None else ""
                label = normalize_age_group(label_raw)
                if not label or not is_age_group(label):
                    continue
                if label in seen_groups:
                    continue
                seen_groups.add(label)
                try:
                    h = float(row[2]) if len(row) > 2 and row[2] else 0
                    m = float(row[3]) if len(row) > 3 and row[3] else 0
                    t = float(row[4]) if len(row) > 4 and row[4] else 0
                    items.append(
                        {
                            "grupo": label,
                            "hombres": round(abs(h), 0),
                            "mujeres": round(abs(m), 0),
                            "total": round(abs(t), 0),
                        }
                    )
                except Exception:
                    pass
            if items:
                cd[key] = items
    # X.8 - Monto promedio vejez por edad
    sk = f'{sec_num}.8'
    if sk in all_data:
        items = []
        seen_groups: set[str] = set()
        for row in all_data[sk]:
            label_raw = str(row[1]) if len(row) > 1 and row[1] is not None else ""
            label = normalize_age_group(label_raw)
            if not label or not is_age_group(label):
                continue
            if label in seen_groups:
                continue
            seen_groups.add(label)
            try:
                h = float(row[2]) if len(row) > 2 and row[2] else 0
                m = float(row[3]) if len(row) > 3 and row[3] else 0
                t = float(row[4]) if len(row) > 4 and row[4] else 0
                items.append(
                    {
                        "grupo": label,
                        "hombres": round(abs(h), 0),
                        "mujeres": round(abs(m), 0),
                        "total": round(abs(t), 0),
                    }
                )
            except Exception:
                pass
        if items:
            cd["monto_vejez_edad"] = items
    summary[caja_name]=cd

def clean(obj):
    if isinstance(obj,dict): return {k:clean(v) for k,v in obj.items()}
    elif isinstance(obj,list): return [clean(v) for v in obj]
    elif isinstance(obj,float): return round(obj,2)
    elif isinstance(obj,int): return obj
    elif obj is None: return 0
    else: return str(obj)

data_json = json.dumps(clean(summary), ensure_ascii=False)

html = '''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Boletín de Seguridad Social 2025 — Dashboard Interactivo | MTESS</title>
<script src="https://cdn.plot.ly/plotly-2.35.0.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;0,9..40,800&family=Source+Serif+4:wght@400;600;700&display=swap" rel="stylesheet">
<link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css" rel="stylesheet">
<style>
:root {
    --mtess-red: #ea2424;
    --mtess-red-dark: #c41e1e;
    --mtess-navy: #1d4354;
    --mtess-navy-deep: #122c38;
    --mtess-cyan: #0693e3;
    --mtess-teal: #34e2e4;
    --bg: #f5f7fa;
    --card: #ffffff;
    --text: #1a2332;
    --text-secondary: #5a6a7e;
    --border: #e3e8ef;
    --shadow-sm: 0 1px 3px rgba(29,67,84,0.08);
    --shadow: 0 4px 12px rgba(29,67,84,0.1);
    --shadow-lg: 0 12px 32px rgba(29,67,84,0.12);
    --radius: 12px;
    --radius-lg: 16px;
}

* { margin:0; padding:0; box-sizing:border-box; }

html { scroll-behavior: smooth; }

:focus-visible {
    outline: 3px solid rgba(6,147,227,0.35);
    outline-offset: 2px;
}

body {
    font-family: 'DM Sans', -apple-system, BlinkMacSystemFont, sans-serif;
    background:
        radial-gradient(1000px 420px at 12% -10%, rgba(52,226,228,0.14), transparent 60%),
        radial-gradient(900px 380px at 92% 8%, rgba(6,147,227,0.12), transparent 55%),
        radial-gradient(1000px 520px at 28% 115%, rgba(234,36,36,0.08), transparent 60%),
        var(--bg);
    background-attachment: fixed;
    color: var(--text);
    line-height: 1.6;
    -webkit-font-smoothing: antialiased;
}

/* ===== TOP BAR ===== */
.topbar {
    background: var(--mtess-navy-deep);
    padding: 8px 32px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    font-size: 12px;
    color: rgba(255,255,255,0.7);
}
.topbar a { color: var(--mtess-teal); text-decoration: none; }

/* ===== HERO ===== */
.hero {
    background: linear-gradient(135deg, var(--mtess-navy-deep) 0%, var(--mtess-navy) 60%, #245e73 100%);
    color: white;
    padding: 0;
    position: relative;
    overflow: hidden;
}
.hero-inner {
    max-width: 1440px;
    margin: 0 auto;
    padding: 44px 40px 40px;
    display: grid;
    grid-template-columns: 1fr auto;
    gap: 40px;
    align-items: center;
    position: relative;
    z-index: 2;
}
.hero::after {
    content: '';
    position: absolute;
    top: -60%;
    right: -15%;
    width: 700px;
    height: 700px;
    background: radial-gradient(circle, rgba(52,226,228,0.08) 0%, transparent 70%);
    border-radius: 50%;
}
.hero-text h1 {
    font-family: 'Source Serif 4', Georgia, serif;
    font-size: clamp(30px, 3.5vw, 46px);
    font-weight: 700;
    line-height: 1.15;
    margin-bottom: 12px;
}
.hero-text h1 span { color: var(--mtess-teal); }
.hero-text p {
    font-size: 16px;
    opacity: 0.8;
    max-width: 580px;
    line-height: 1.7;
}
.hero-badge {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    background: var(--mtess-red);
    padding: 5px 14px;
    border-radius: 4px;
    font-size: 11px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1.5px;
    margin-bottom: 16px;
}
.hero-kpis {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: 12px;
}
.hero-kpi {
    background: rgba(255,255,255,0.08);
    backdrop-filter: blur(12px);
    border: 1px solid rgba(255,255,255,0.12);
    border-radius: var(--radius);
    padding: 18px 20px;
    text-align: center;
    min-width: 170px;
    transition: transform 0.22s ease, box-shadow 0.22s ease, border-color 0.22s ease;
}
.hero-kpi:hover {
    transform: translateY(-2px);
    box-shadow: 0 14px 28px rgba(0,0,0,0.18);
    border-color: rgba(52,226,228,0.28);
}
.hero-kpi .num {
    font-size: 28px;
    font-weight: 800;
    color: var(--mtess-teal);
    letter-spacing: -0.5px;
}
.hero-kpi .lbl {
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    opacity: 0.65;
    margin-top: 2px;
}
.red-stripe {
    height: 4px;
    background: linear-gradient(90deg, var(--mtess-red) 0%, var(--mtess-red-dark) 50%, var(--mtess-red) 100%);
}

/* ===== NAV ===== */
.nav {
    background: rgba(255,255,255,0.86);
    backdrop-filter: blur(14px);
    -webkit-backdrop-filter: blur(14px);
    border-bottom: 1px solid rgba(227,232,239,0.9);
    position: sticky;
    top: 0;
    z-index: 100;
    box-shadow: var(--shadow-sm);
}
.nav-inner {
    max-width: 1440px;
    margin: 0 auto;
    display: flex;
    align-items: center;
    padding: 0 40px;
    overflow-x: auto;
    scrollbar-width: none;
    -ms-overflow-style: none;
}
.nav-inner::-webkit-scrollbar { display:none; }
.nav-btn {
    padding: 16px 20px;
    font-size: 13px;
    font-weight: 600;
    color: var(--text-secondary);
    background: none;
    border: none;
    cursor: pointer;
    white-space: nowrap;
    border-bottom: 3px solid transparent;
    transition: all 0.2s;
    display: flex;
    align-items: center;
    gap: 6px;
}
.nav-btn i { font-size: 14px; }
.nav-btn:hover { color: var(--mtess-red); }
.nav-btn.active {
    color: var(--mtess-red);
    border-bottom-color: var(--mtess-red);
    background: linear-gradient(180deg, rgba(234,36,36,0.10), transparent);
}

/* ===== MAIN ===== */
.main {
    max-width: 1440px;
    margin: 0 auto;
    padding: 28px 40px 60px;
}

.section { display:none; animation: fadeUp 0.35s ease; }
.section.active { display:block; }
@keyframes fadeUp {
    from { opacity:0; transform:translateY(12px); }
    to { opacity:1; transform:translateY(0); }
}

.sec-header {
    margin-bottom: 28px;
    padding-bottom: 16px;
    border-bottom: 2px solid var(--border);
}
.sec-header h2 {
    font-family: 'Source Serif 4', Georgia, serif;
    font-size: 26px;
    font-weight: 700;
    color: var(--mtess-navy);
}
.sec-header p {
    font-size: 14px;
    color: var(--text-secondary);
    margin-top: 4px;
}

/* ===== KPI CARDS ===== */
.kpi-row {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 16px;
    margin-bottom: 28px;
}
.kpi {
    background: var(--card);
    border-radius: var(--radius);
    padding: 22px 24px;
    box-shadow: var(--shadow-sm);
    border: 1px solid var(--border);
    border-left: 4px solid var(--mtess-navy);
    transition: transform 0.2s, box-shadow 0.2s;
}
.kpi:hover { transform: translateY(-2px); box-shadow: var(--shadow); }
.kpi.red { border-left-color: var(--mtess-red); }
.kpi.teal { border-left-color: var(--mtess-teal); }
.kpi.cyan { border-left-color: var(--mtess-cyan); }
.kpi .kpi-icon {
    width: 40px; height: 40px; border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 18px; margin-bottom: 12px;
}
.kpi .kpi-icon.navy { background: rgba(29,67,84,0.1); color: var(--mtess-navy); }
.kpi .kpi-icon.red { background: rgba(234,36,36,0.1); color: var(--mtess-red); }
.kpi .kpi-icon.teal { background: rgba(52,226,228,0.1); color: #0d9e9f; }
.kpi .kpi-icon.cyan { background: rgba(6,147,227,0.08); color: var(--mtess-cyan); }
.kpi .kpi-lbl {
    font-size: 11px; font-weight: 700; text-transform: uppercase;
    letter-spacing: 0.8px; color: var(--text-secondary); margin-bottom: 4px;
}
.kpi .kpi-val {
    font-size: 26px; font-weight: 800; color: var(--text); letter-spacing: -0.5px;
}
.kpi .kpi-sub { font-size: 12px; font-weight: 600; margin-top: 4px; }
.kpi .kpi-sub.up { color: #0d9e9f; }
.kpi .kpi-sub.down { color: var(--mtess-red); }

/* ===== CHART CARDS ===== */
.grid { display: grid; gap: 20px; margin-bottom: 24px; }
.g2 { grid-template-columns: repeat(2, 1fr); }
.g3 { grid-template-columns: repeat(3, 1fr); }
.g1 { grid-template-columns: 1fr; }
@media (max-width:900px) { .g2,.g3 { grid-template-columns:1fr; } }

.card {
    position: relative;
    background: var(--card);
    border-radius: var(--radius-lg);
    box-shadow: var(--shadow-sm);
    border: 1px solid rgba(227,232,239,0.9);
    overflow: hidden;
    transition: transform 0.22s ease, box-shadow 0.22s ease, border-color 0.22s ease;
}
.card::before {
    content: '';
    position: absolute;
    inset: 0;
    background:
        radial-gradient(700px 260px at 10% 0%, rgba(52,226,228,0.14), transparent 60%),
        radial-gradient(700px 260px at 110% 0%, rgba(6,147,227,0.10), transparent 55%),
        radial-gradient(800px 320px at 30% 120%, rgba(234,36,36,0.08), transparent 65%);
    opacity: 0;
    transition: opacity 0.22s ease;
    pointer-events: none;
}
.card > * { position: relative; z-index: 1; }
.card:hover {
    transform: translateY(-2px);
    box-shadow: var(--shadow-lg);
    border-color: rgba(6,147,227,0.22);
}
.card:hover::before { opacity: 1; }
.card-head {
    padding: 18px 24px 0;
    display: flex;
    align-items: flex-start;
    justify-content: space-between;
}
.card-head h3 {
    font-size: 15px; font-weight: 700; color: var(--text);
}
.card-head .tag {
    font-size: 10px; font-weight: 700; text-transform: uppercase;
    letter-spacing: 1px; padding: 3px 10px; border-radius: 4px;
    background: rgba(29,67,84,0.08); color: var(--mtess-navy);
}
.card-head .tag.red { background: rgba(234,36,36,0.08); color: var(--mtess-red); }
.card-sub { padding: 0 24px; font-size: 12px; color: var(--text-secondary); margin-top: 2px; }
.card-body { padding: 8px 12px 14px; }
.span2 { grid-column: span 2; }
@media (max-width:900px) { .span2 { grid-column: span 1; } }

/* ===== CAJA SELECTOR ===== */
.pill-row { display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 22px; }
.pill {
    padding: 8px 20px; border: 2px solid var(--border); border-radius: 100px;
    background: white; font-size: 13px; font-weight: 600;
    color: var(--text-secondary); cursor: pointer; transition: all 0.2s;
}
.pill:hover { border-color: var(--mtess-red); color: var(--mtess-red); }
.pill.on {
    background: linear-gradient(135deg, var(--mtess-red) 0%, #ff4d4d 100%);
    border-color: rgba(234,36,36,0.95);
    color: white;
    box-shadow: 0 10px 22px rgba(234,36,36,0.22);
}

/* ===== TABLE ===== */
.tbl { width:100%; border-collapse:separate; border-spacing:0; font-size:13px; }
.tbl thead th {
    background: var(--mtess-navy); color: white;
    padding: 12px 16px; text-align: right;
    font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.6px;
}
.tbl thead th:first-child { text-align:left; border-radius:8px 0 0 0; }
.tbl thead th:last-child { border-radius:0 8px 0 0; }
.tbl td {
    padding: 11px 16px; text-align:right; border-bottom:1px solid var(--border);
}
.tbl td:first-child { text-align:left; font-weight:600; color: var(--mtess-navy); }
.tbl tr:hover td { background: rgba(6,147,227,0.04); }
.tbl tr:last-child td { font-weight:700; background: rgba(29,67,84,0.05); }

/* ===== INSIGHTS ===== */
.ins-list { padding-left: 18px; margin: 12px 0 6px; }
.ins-list li { margin: 8px 0; color: var(--text-secondary); }
.ins-list strong { color: var(--text); }

/* ===== FOOTER ===== */
.footer {
    background: var(--mtess-navy-deep); color: rgba(255,255,255,0.6);
    padding: 36px 40px; text-align: center; font-size: 13px;
}
.footer strong { color: white; }
.footer .red { color: var(--mtess-red); }

/* ===== FAB ===== */
.fab {
    position: fixed; bottom: 24px; right: 24px;
    width: 52px; height: 52px; border-radius: 50%;
    background: var(--mtess-red); color: white; border: none;
    cursor: pointer; box-shadow: var(--shadow-lg);
    font-size: 18px; display: flex; align-items: center; justify-content: center;
    z-index: 999; transition: transform 0.2s;
}
.fab:hover { transform: scale(1.1); }

@media print {
    .nav,.fab,.topbar { display:none; }
    .section { display:block!important; page-break-after:always; }
}

@media (max-width:768px) {
    .hero-inner { grid-template-columns:1fr; padding: 32px 20px; }
    .hero-kpis { grid-template-columns: repeat(2,1fr); }
    .main { padding: 16px; }
    .kpi-row { grid-template-columns: repeat(2,1fr); }
}
</style>
</head>
<body>

<div class="topbar">
    <span><i class="fas fa-landmark"></i> &nbsp;Ministerio de Trabajo, Empleo y Seguridad Social — República del Paraguay</span>
    <a href="https://www.mtess.gov.py" target="_blank">www.mtess.gov.py</a>
</div>

<div class="hero">
    <div class="hero-inner">
        <div class="hero-text">
            <div class="hero-badge"><i class="fas fa-shield-halved"></i> Dirección General de Seguridad Social</div>
            <h1>Boletín de<br>Seguridad Social <span>2025</span></h1>
            <p>Anexo Estadístico — Dashboard interactivo con datos consolidados del sistema previsional del Paraguay, período 2020–2024.</p>
        </div>
        <div class="hero-kpis" id="heroKpis"></div>
    </div>
</div>
<div class="red-stripe"></div>

<div class="nav"><div class="nav-inner" id="navBar"></div></div>
<div class="main" id="mainContent"></div>

<div class="footer">
    <p><strong>Boletín de Seguridad Social 2025</strong> — Anexo Estadístico</p>
    <p style="margin-top:4px">Fuente: <strong class="red">MTESS</strong> – Dirección General de Seguridad Social, a partir de datos de las cajas previsionales.</p>
    <p style="margin-top:8px;opacity:0.5">Los datos presentados corresponden a diciembre de cada año. &copy; 2025</p>
</div>

<button class="fab" onclick="window.print()" title="Imprimir / Exportar PDF"><i class="fas fa-print"></i></button>

<script>
const D = ''' + data_json + ''';

// MTESS palette
const C = {
    navy:'#1d4354', navyDeep:'#122c38', red:'#ea2424', redDark:'#c41e1e',
    cyan:'#0693e3', teal:'#0d9e9f', tealLight:'#34e2e4',
    blue:'#2c7be5', blueLight:'#6cb4ee',
    amber:'#f6a609', orange:'#e8590c',
    green:'#0ca678', greenLight:'#69db7c',
    purple:'#7c3aed', pink:'#d6336c',
    gray:'#adb5bd', grayLight:'#dee2e6'
};
const CAJA_C = {
    'IPS':C.navy, 'Caja Fiscal':C.red, 'Caja ANDE':C.teal,
    'Caja Bancaria':C.purple, 'Caja Municipal':C.orange,
    'Caja Ferroviaria':C.blue, 'CAJUBI':C.pink, 'Caja Parlamentaria':C.amber
};
const CAJA_C2 = {
    'IPS':C.cyan, 'Caja Fiscal':'#ff6b6b', 'Caja ANDE':C.tealLight,
    'Caja Bancaria':'#b197fc', 'Caja Municipal':'#ffa94d',
    'Caja Ferroviaria':C.blueLight, 'CAJUBI':'#f783ac', 'Caja Parlamentaria':'#ffd43b'
};
const YRS = ['2020','2021','2022','2023','2024'];

function fmt(n){ return n==null||n===0?'0':Number(n).toLocaleString('es-PY'); }
function fmtGs(n){ return n==null||n===0?'—':'Gs. '+Math.round(Number(n)).toLocaleString('es-PY'); }
function fmtM(n){ const v=Number(n); if(v>=1e6) return (v/1e6).toFixed(1)+'M'; if(v>=1e3) return (v/1e3).toFixed(0)+'K'; return fmt(n); }
function pct(o,c){ if(!o||o===0) return {t:'',c:''}; const p=((c-o)/o*100).toFixed(1); return {t:(p>0?'+':'')+p+'%', c:p>=0?'up':'down'}; }
function safeDiv(n,d){ return d? (n/d) : 0; }

function ageLower(label){
    const s=String(label||'').toLowerCase();
    const m=s.match(/(\d+)\s*a\s*\d+/);
    if(m) return Number(m[1]);
    const m2=s.match(/(\d+)\s*(?:y\s*m[aá]s|\+)/);
    if(m2) return Number(m2[1]);
    return null;
}
function ageMid(label){
    const s=String(label||'').toLowerCase();
    const m=s.match(/(\d+)\s*a\s*(\d+)/);
    if(m) return (Number(m[1])+Number(m[2]))/2;
    const lb=ageLower(s);
    if(lb!=null) return lb + 2.5; // sup. tramo abierto ~5 años
    return null;
}
function ageStats(items){
    let tot=0, sum=0, tot55=0;
    (items||[]).forEach(e=>{
        const t=Number(e.total || (Number(e.hombres||0)+Number(e.mujeres||0)) || 0);
        const mid=ageMid(e.grupo);
        const lb=ageLower(e.grupo);
        if(t>0){
            tot += t;
            if(mid!=null) sum += mid*t;
            if(lb!=null && lb>=55) tot55 += t;
        }
    });
    return {total: tot, avg: safeDiv(sum, tot), share55: safeDiv(tot55, tot)};
}
function payGapActivos(c){
    const counts=c?.activos_edad;
    const sal=c?.salario_edad;
    if(!counts?.length || !sal?.length) return null;
    const salBy={};
    sal.forEach(e=>{ if(e?.grupo) salBy[e.grupo]=e; });
    let sumH=0, sumM=0, cntH=0, cntM=0;
    counts.forEach(e=>{
        const s=salBy[e.grupo];
        if(!s) return;
        const hC=Number(e.hombres||0), mC=Number(e.mujeres||0);
        const hS=Number(s.hombres||0), mS=Number(s.mujeres||0);
        if(hC>0 && hS>0){ sumH += hC*hS; cntH += hC; }
        if(mC>0 && mS>0){ sumM += mC*mS; cntM += mC; }
    });
    const avgH=safeDiv(sumH, cntH);
    const avgM=safeDiv(sumM, cntM);
    const ratio=avgH>0? (avgM/avgH) : 0;
    return {avgH, avgM, ratio};
}

// Layout defaults
const LO = {
    margin:{t:28,b:48,l:64,r:24}, height:380,
    paper_bgcolor:'rgba(0,0,0,0)', plot_bgcolor:'rgba(0,0,0,0)',
    font:{family:'DM Sans, sans-serif', size:12, color:'#5a6a7e'},
    xaxis:{gridcolor:'#eef1f5', gridwidth:1, zeroline:false, tickfont:{size:11}},
    yaxis:{gridcolor:'#eef1f5', gridwidth:1, zeroline:false, tickfont:{size:11}},
    legend:{orientation:'h', y:1.12, x:0.5, xanchor:'center', font:{size:12}},
    hoverlabel:{font:{family:'DM Sans',size:13}, bordercolor:'transparent'},
    modebar:{bgcolor:'transparent', color:'#adb5bd', activecolor:C.red}
};
const CFG = {responsive:true, displaylogo:false, modeBarButtonsToRemove:['lasso2d','select2d']};

// Totals
let totA24=0, totP24=0, totA23=0, totP23=0;
Object.values(D).forEach(c=>{
    totA24+=Number(c.activos?.['2024']||0); totP24+=Number(c.pasivos?.['2024']||0);
    totA23+=Number(c.activos?.['2023']||0); totP23+=Number(c.pasivos?.['2023']||0);
});
const ipsA24 = Number(D['IPS']?.activos?.['2024'] || 0);
const ipsP24 = Number(D['IPS']?.pasivos?.['2024'] || 0);
const ipsAShare = totA24 > 0 ? (ipsA24 / totA24 * 100) : 0;
const ipsPShare = totP24 > 0 ? (ipsP24 / totP24 * 100) : 0;

// HERO KPIs
document.getElementById('heroKpis').innerHTML = `
    <div class="hero-kpi"><div class="num">${fmtM(totA24)}</div><div class="lbl">Cotizantes Activos</div></div>
    <div class="hero-kpi"><div class="num">${fmtM(totP24)}</div><div class="lbl">Jubilados/Pensionados</div></div>
    <div class="hero-kpi"><div class="num">${(totA24/totP24).toFixed(1)}</div><div class="lbl">Relación A/P Global</div></div>
    <div class="hero-kpi"><div class="num">8</div><div class="lbl">Cajas Previsionales</div></div>
`;

// SECTIONS CONFIG
const SECS = [
    {id:'resumen', label:'Resumen', icon:'fa-chart-pie'},
    {id:'evolucion', label:'Evolución', icon:'fa-chart-line'},
    {id:'cajas', label:'Por Caja', icon:'fa-building-columns'},
    {id:'genero', label:'Género', icon:'fa-venus-mars'},
    {id:'edades', label:'Pirámides', icon:'fa-people-group'},
    {id:'salarios', label:'Salarios', icon:'fa-sack-dollar'},
    {id:'tabla', label:'Comparativa', icon:'fa-table-cells'},
];

// NAV
const nav = document.getElementById('navBar');
SECS.forEach((s,i)=>{
    const b=document.createElement('button');
    b.className='nav-btn'+(i===0?' active':'');
    b.innerHTML=`<i class="fas ${s.icon}"></i>${s.label}`;
    b.onclick=()=>showSec(s.id);
    nav.appendChild(b);
});
function showSec(id){
    document.querySelectorAll('.section').forEach(s=>s.classList.remove('active'));
    document.querySelectorAll('.nav-btn').forEach(b=>b.classList.remove('active'));
    document.getElementById('s-'+id).classList.add('active');
    nav.children[SECS.findIndex(s=>s.id===id)].classList.add('active');
    window.dispatchEvent(new Event('resize')); // fix plotly resize
}

const M = document.getElementById('mainContent');

// ======================== SEC: RESUMEN ========================
!function(){
    const s=document.createElement('div'); s.id='s-resumen'; s.className='section active';
    const chA=pct(totA23,totA24), chP=pct(totP23,totP24);
    s.innerHTML=`
    <div class="sec-header"><h2>Resumen General del Sistema Previsional</h2><p>Visión consolidada de todas las cajas — Datos a diciembre 2024</p></div>
	    <div class="kpi-row">
	        <div class="kpi"><div class="kpi-icon navy"><i class="fas fa-user-tie"></i></div><div class="kpi-lbl">Cotizantes Activos</div><div class="kpi-val">${fmt(totA24)}</div><div class="kpi-sub ${chA.c}"><i class="fas fa-arrow-trend-${chA.c==='up'?'up':'down'}"></i> ${chA.t} vs 2023</div></div>
	        <div class="kpi red"><div class="kpi-icon red"><i class="fas fa-user-clock"></i></div><div class="kpi-lbl">Jubilados/Pensionados</div><div class="kpi-val">${fmt(totP24)}</div><div class="kpi-sub ${chP.c}"><i class="fas fa-arrow-trend-${chP.c==='up'?'up':'down'}"></i> ${chP.t} vs 2023</div></div>
	        <div class="kpi teal"><div class="kpi-icon teal"><i class="fas fa-scale-balanced"></i></div><div class="kpi-lbl">Relación Activos/Pasivos</div><div class="kpi-val">${(totA24/totP24).toFixed(2)}</div><div class="kpi-sub">activos por cada pasivo</div></div>
	        <div class="kpi cyan"><div class="kpi-icon cyan"><i class="fas fa-users"></i></div><div class="kpi-lbl">Beneficiarios Totales</div><div class="kpi-val">${fmt(totA24+totP24)}</div><div class="kpi-sub">en 8 cajas previsionales</div></div>
	        <div class="kpi teal"><div class="kpi-icon teal"><i class="fas fa-building-columns"></i></div><div class="kpi-lbl">IPS — Activos</div><div class="kpi-val">${fmt(ipsA24)}</div><div class="kpi-sub">${ipsAShare.toFixed(1)}% del total de activos</div></div>
	        <div class="kpi red"><div class="kpi-icon red"><i class="fas fa-building-columns"></i></div><div class="kpi-lbl">IPS — Pasivos</div><div class="kpi-val">${fmt(ipsP24)}</div><div class="kpi-sub">${ipsPShare.toFixed(1)}% del total de pasivos</div></div>
	    </div>
	    <div class="grid g2">
	        <div class="card"><div class="card-head"><h3>Distribución de Cotizantes Activos</h3><div class="tag">2024</div></div><div class="card-body" id="r-pie1"></div></div>
	        <div class="card"><div class="card-head"><h3>Distribución de Jubilados/Pensionados</h3><div class="tag red">2024</div></div><div class="card-body" id="r-pie2"></div></div>
		    </div>
			    <div class="grid g1">
			        <div class="card"><div class="card-head"><h3>Activos vs Pasivos por Caja Previsional</h3><div class="tag">2024</div></div><div class="card-sub">Comparación absoluta entre cotizantes y beneficiarios</div><div class="card-body" id="r-bar1"></div></div>
			    </div>
			    <div class="grid g2">
			        <div class="card"><div class="card-head"><h3>IPS: Participación en el Sistema</h3><div class="tag">2024</div></div><div class="card-sub">Cuotas de IPS sobre el total (activos y pasivos)</div><div class="card-body" id="r-ips"></div></div>
			        <div class="card"><div class="card-head"><h3>Hallazgos Rápidos</h3><div class="tag red">2024</div></div><div class="card-sub">Indicadores destacados generados automáticamente</div><div class="card-body" id="r-ins"></div></div>
			    </div>`;
    M.appendChild(s);

    setTimeout(()=>{
        const cn=Object.keys(D), a24=cn.map(c=>Number(D[c].activos?.['2024']||0)), p24=cn.map(c=>Number(D[c].pasivos?.['2024']||0));
        const cc=cn.map(c=>CAJA_C[c]), cc2=cn.map(c=>CAJA_C2[c]);

        const pieL={margin:{t:8,b:8,l:8,r:8},height:340,paper_bgcolor:'rgba(0,0,0,0)',font:{family:'DM Sans',size:12},showlegend:false};
        Plotly.newPlot('r-pie1',[{values:a24,labels:cn,type:'pie',hole:0.52,marker:{colors:cc,line:{color:'white',width:2}},textinfo:'percent',textposition:'inside',textfont:{size:12,color:'white'},hovertemplate:'<b>%{label}</b><br>%{value:,.0f} activos<br>%{percent}<extra></extra>',sort:false,direction:'clockwise',pull:cn.map((_,i)=>i===0?0.03:0)}],{...pieL,annotations:[{text:'<b>Activos</b>',showarrow:false,font:{size:14,color:C.navy}}]},CFG);
        Plotly.newPlot('r-pie2',[{values:p24,labels:cn,type:'pie',hole:0.52,marker:{colors:cc,line:{color:'white',width:2}},textinfo:'percent',textposition:'inside',textfont:{size:12,color:'white'},hovertemplate:'<b>%{label}</b><br>%{value:,.0f} pasivos<br>%{percent}<extra></extra>',sort:false,direction:'clockwise'}],{...pieL,annotations:[{text:'<b>Pasivos</b>',showarrow:false,font:{size:14,color:C.red}}]},CFG);

	        Plotly.newPlot('r-bar1',[
	            {x:cn,y:a24,name:'Cotizantes Activos',type:'bar',marker:{color:cn.map(c=>CAJA_C[c]),line:{width:0}},hovertemplate:'<b>%{x}</b><br>Activos: %{y:,.0f}<extra></extra>',text:a24.map(v=>fmtM(v)),textposition:'outside',textfont:{size:11,color:C.navy}},
	            {x:cn,y:p24,name:'Jubilados/Pensionados',type:'bar',marker:{color:cn.map(c=>CAJA_C2[c]),line:{width:0}},hovertemplate:'<b>%{x}</b><br>Pasivos: %{y:,.0f}<extra></extra>',text:p24.map(v=>fmtM(v)),textposition:'outside',textfont:{size:11,color:C.red}}
	        ],{...LO,barmode:'group',bargap:0.25,bargroupgap:0.08,height:420,margin:{...LO.margin,b:80,t:36}},CFG);

	        // IPS: cuotas sobre el total del sistema
	        const restA = Math.max(0, totA24 - ipsA24);
	        const restP = Math.max(0, totP24 - ipsP24);
	        Plotly.newPlot('r-ips',[
	            {values:[ipsA24,restA],labels:['IPS','Resto'],type:'pie',hole:0.55,domain:{x:[0,0.48],y:[0,1]},
	             marker:{colors:[C.navy,'rgba(29,67,84,0.14)'],line:{color:'white',width:2}},
	             textinfo:'percent',textposition:'inside',textfont:{size:12,color:'white'},
	             hovertemplate:'<b>%{label}</b><br>%{value:,.0f}<br>%{percent}<extra></extra>',sort:false},
	            {values:[ipsP24,restP],labels:['IPS','Resto'],type:'pie',hole:0.55,domain:{x:[0.52,1],y:[0,1]},
	             marker:{colors:[C.red,'rgba(234,36,36,0.14)'],line:{color:'white',width:2}},
	             textinfo:'percent',textposition:'inside',textfont:{size:12,color:'white'},
	             hovertemplate:'<b>%{label}</b><br>%{value:,.0f}<br>%{percent}<extra></extra>',sort:false}
	        ],{...pieL,annotations:[
	            {x:0.24,y:0.5,text:`<b>Activos</b><br>${fmt(ipsA24)}<br><span style="color:${C.gray}">${ipsAShare.toFixed(1)}%</span>`,showarrow:false,font:{size:12,color:C.navy}},
	            {x:0.76,y:0.5,text:`<b>Pasivos</b><br>${fmt(ipsP24)}<br><span style="color:${C.gray}">${ipsPShare.toFixed(1)}%</span>`,showarrow:false,font:{size:12,color:C.red}}
	        ]},CFG);

	        // Hallazgos automáticos
	        const topAct = cn.map((c,i)=>({c,v:a24[i]})).sort((a,b)=>b.v-a.v)[0];
	        const topFem = cn.map(c=>{
	            const h=Number(D[c].por_sexo?.activos?.Hombres?.['2024']||0);
	            const m=Number(D[c].por_sexo?.activos?.Mujeres?.['2024']||0);
	            const t=h+m;
	            return t?{c,p:m/t*100}:null;
	        }).filter(Boolean).sort((a,b)=>b.p-a.p)[0];
	        const oldest = cn.map(c=>{
	            const st=ageStats(D[c].activos_edad);
	            return st.total?{c,avg:st.avg,share55:st.share55*100}:null;
	        }).filter(Boolean).sort((a,b)=>b.avg-a.avg)[0];
	        const worstRatio = cn.map((c,i)=>({c,r:(p24[i]>0)?(a24[i]/p24[i]):null})).filter(o=>o.r!=null).sort((a,b)=>a.r-b.r)[0];
	        const bestRR = cn.map(c=>{const r=Number(D[c].salarios?.['2024']?.ratio||0); return r?{c,r:r*100}:null;}).filter(Boolean).sort((a,b)=>b.r-a.r)[0];

	        const insights=[];
	        if(topAct) insights.push(`<strong>${topAct.c}</strong> concentra ${((topAct.v/totA24)*100).toFixed(1)}% de los activos (${fmt(topAct.v)}).`);
	        if(topFem) insights.push(`Mayor participación femenina en activos: <strong>${topFem.c}</strong> (${topFem.p.toFixed(1)}% mujeres).`);
	        if(oldest) insights.push(`Estructura de activos más envejecida: <strong>${oldest.c}</strong> (edad promedio estimada ${oldest.avg.toFixed(1)}; ${oldest.share55.toFixed(1)}% tiene 55+).`);
	        if(worstRatio) insights.push(`Relación Activos/Pasivos más baja (2024): <strong>${worstRatio.c}</strong> (${worstRatio.r.toFixed(2)}).`);
	        if(bestRR) insights.push(`Mayor tasa de reemplazo promedio (2024): <strong>${bestRR.c}</strong> (${bestRR.r.toFixed(1)}%).`);
	        document.getElementById('r-ins').innerHTML = `<div style="padding: 10px 16px 6px"><ul class="ins-list">${insights.map(t=>`<li>${t}</li>`).join('')}</ul></div>`;
	    },50);
	}();

// ======================== SEC: EVOLUCION ========================
!function(){
    const s=document.createElement('div'); s.id='s-evolucion'; s.className='section';
    s.innerHTML=`
    <div class="sec-header"><h2>Evolución Temporal 2020–2024</h2><p>Tendencias del sistema previsional en los últimos cinco años</p></div>
    <div class="grid g1">
        <div class="card"><div class="card-head"><h3>Evolución Total del Sistema: Activos y Pasivos</h3><div class="tag">2020–2024</div></div><div class="card-sub">Suma consolidada de todas las cajas</div><div class="card-body" id="e-total"></div></div>
    </div>
    <div class="grid g2">
        <div class="card"><div class="card-head"><h3>Cotizantes Activos por Caja</h3></div><div class="card-body" id="e-act"></div></div>
        <div class="card"><div class="card-head"><h3>Jubilados/Pensionados por Caja</h3></div><div class="card-body" id="e-pas"></div></div>
    </div>
    <div class="grid g1">
        <div class="card"><div class="card-head"><h3>Relación Activos/Pasivos — Indicador de Sostenibilidad</h3><div class="tag red">Crítico</div></div><div class="card-sub">La línea roja punteada indica el umbral crítico de 2 activos por pasivo</div><div class="card-body" id="e-rel"></div></div>
    </div>`;
    M.appendChild(s);

    setTimeout(()=>{
        const tA=YRS.map(y=>Object.values(D).reduce((s,c)=>s+Number(c.activos?.[y]||0),0));
        const tP=YRS.map(y=>Object.values(D).reduce((s,c)=>s+Number(c.pasivos?.[y]||0),0));
        Plotly.newPlot('e-total',[
            {x:YRS,y:tA,name:'Activos',fill:'tozeroy',fillcolor:'rgba(29,67,84,0.08)',line:{color:C.navy,width:3,shape:'spline',smoothing:1.15},marker:{size:10,color:C.navy},mode:'lines+markers',hovertemplate:'Activos: %{y:,.0f}<extra></extra>'},
            {x:YRS,y:tP,name:'Pasivos',fill:'tozeroy',fillcolor:'rgba(234,36,36,0.06)',line:{color:C.red,width:3,shape:'spline',smoothing:1.15},marker:{size:10,color:C.red},mode:'lines+markers',hovertemplate:'Pasivos: %{y:,.0f}<extra></extra>'}
        ],{...LO,height:360},CFG);

        const mkTrace=(key,dash)=>Object.entries(D).map(([n,c])=>({x:YRS,y:YRS.map(y=>Number(c[key]?.[y]||0)),name:n,mode:'lines+markers',line:{color:CAJA_C[n],width:2.5,shape:'spline',smoothing:1.15,dash},marker:{size:7,color:CAJA_C[n]}}));
        Plotly.newPlot('e-act',mkTrace('activos'),{...LO,legend:{...LO.legend,font:{size:10}}},CFG);
        Plotly.newPlot('e-pas',mkTrace('pasivos'),{...LO,legend:{...LO.legend,font:{size:10}}},CFG);

        const relT=Object.entries(D).filter(([,c])=>c.relacion).map(([n,c])=>({x:YRS,y:YRS.map(y=>Number(c.relacion?.[y]||0)),name:n,mode:'lines+markers',line:{color:CAJA_C[n],width:2.5,shape:'spline',smoothing:1.15},marker:{size:7}}));
        Plotly.newPlot('e-rel',relT,{...LO,height:400,yaxis:{...LO.yaxis,title:{text:'Activos por cada Pasivo',font:{size:13}}},shapes:[{type:'line',x0:YRS[0],x1:YRS[4],y0:2,y1:2,line:{color:C.red,width:2,dash:'dash'}}],annotations:[{x:YRS[4],y:2.15,text:'<b>Umbral crítico</b>',showarrow:false,font:{size:11,color:C.red}}]},CFG);
    },100);
}();

// ======================== SEC: POR CAJA ========================
!function(){
    const s=document.createElement('div'); s.id='s-cajas'; s.className='section';
    const cn=Object.keys(D);
    s.innerHTML=`
    <div class="sec-header"><h2>Análisis Detallado por Caja</h2><p>Seleccione una caja previsional</p></div>
    <div class="pill-row">${cn.map((c,i)=>`<button class="pill${i===0?' on':''}" onclick="selCaja('${c}',this)">${c}</button>`).join('')}</div>
    <div id="caja-det"></div>`;
    M.appendChild(s);

    window.selCaja=function(name,btn){
        document.querySelectorAll('#s-cajas .pill').forEach(b=>b.classList.remove('on'));
        if(btn) btn.classList.add('on');
        const c=D[name], det=document.getElementById('caja-det');
        const a24=Number(c.activos?.['2024']||0), p24=Number(c.pasivos?.['2024']||0);
        const rel=p24>0?(a24/p24).toFixed(2):'N/A';
        const ch=pct(Number(c.activos?.['2023']||0),a24);

        det.innerHTML=`
        <div class="kpi-row">
            <div class="kpi"><div class="kpi-icon navy"><i class="fas fa-users"></i></div><div class="kpi-lbl">Activos 2024</div><div class="kpi-val">${fmt(a24)}</div><div class="kpi-sub ${ch.c}">${ch.t} vs 2023</div></div>
            <div class="kpi red"><div class="kpi-icon red"><i class="fas fa-user-clock"></i></div><div class="kpi-lbl">Pasivos 2024</div><div class="kpi-val">${fmt(p24)}</div></div>
            <div class="kpi teal"><div class="kpi-icon teal"><i class="fas fa-scale-balanced"></i></div><div class="kpi-lbl">Relación A/P</div><div class="kpi-val">${rel}</div></div>
            ${c.salarios?.['2024']?`<div class="kpi cyan"><div class="kpi-icon cyan"><i class="fas fa-money-bill-wave"></i></div><div class="kpi-lbl">Salario Promedio</div><div class="kpi-val" style="font-size:18px">${fmtGs(c.salarios['2024'].salario_activo)}</div></div>`:''}
	        </div>
	        <div class="grid g2">
	            <div class="card"><div class="card-head"><h3>Evolución Activos y Pasivos</h3></div><div class="card-body" id="cd-evo"></div></div>
	            <div class="card"><div class="card-head"><h3>Pasivos por Tipo de Beneficio</h3><div class="tag">2024</div></div><div class="card-body" id="cd-tipo"></div></div>
	        </div>
	        ${c.activos_edad?.length?`<div class="grid g1"><div class="card"><div class="card-head"><h3>Cotizantes Activos por Grupo de Edad y Sexo</h3></div><div class="card-body" id="cd-age"></div></div></div>`:''}
	        ${c.salario_edad?.length?`<div class="grid g1"><div class="card"><div class="card-head"><h3>Salario Promedio del Cotizante Activo según Edad y Sexo</h3><div class="tag">Guaraníes</div></div><div class="card-body" id="cd-salEdad"></div></div></div>`:''}
	        ${c.monto_vejez_edad?.length?`<div class="grid g1"><div class="card"><div class="card-head"><h3>Monto Promedio Jubilación por Vejez según Edad y Sexo</h3><div class="tag">Guaraníes</div></div><div class="card-body" id="cd-montoVejez"></div></div></div>`:''}`;

        // Evo chart
        Plotly.newPlot('cd-evo',[
            {x:YRS,y:YRS.map(y=>Number(c.activos?.[y]||0)),name:'Activos',type:'bar',marker:{color:C.navy},hovertemplate:'%{y:,.0f}<extra>Activos</extra>',text:YRS.map(y=>fmtM(c.activos?.[y]||0)),textposition:'outside',textfont:{size:10}},
            {x:YRS,y:YRS.map(y=>Number(c.pasivos?.[y]||0)),name:'Pasivos',type:'bar',marker:{color:C.red},hovertemplate:'%{y:,.0f}<extra>Pasivos</extra>',text:YRS.map(y=>fmtM(c.pasivos?.[y]||0)),textposition:'outside',textfont:{size:10}}
        ],{...LO,barmode:'group',bargap:0.3},CFG);

        // Tipo pie
        if(c.jubilados_tipo){
            const tp=Object.entries(c.jubilados_tipo).filter(([k])=>k!=='Total');
            const tpC=[C.navy,C.amber,C.red,C.purple];
            Plotly.newPlot('cd-tipo',[{
                values:tp.map(([,v])=>Number(v['2024']||0)),labels:tp.map(([k])=>k),
                type:'pie',hole:0.5,marker:{colors:tpC,line:{color:'white',width:2}},
                textinfo:'percent+label',textposition:'inside',textfont:{size:11,color:'white'},
                hovertemplate:'<b>%{label}</b><br>%{value:,.0f}<br>%{percent}<extra></extra>'
            }],{margin:{t:8,b:8,l:8,r:8},height:LO.height,paper_bgcolor:'rgba(0,0,0,0)',font:{family:'DM Sans'},showlegend:false},CFG);
        }

	        // Age pyramid activos
	        if(c.activos_edad?.length && document.getElementById('cd-age')){
	            const g=c.activos_edad.map(e=>e.grupo);
	            Plotly.newPlot('cd-age',[
	                {x:g,y:c.activos_edad.map(e=>e.hombres),name:'Hombres',type:'bar',marker:{color:C.navy,line:{width:0}},hovertemplate:'Hombres: %{y:,.0f}<extra></extra>'},
	                {x:g,y:c.activos_edad.map(e=>e.mujeres),name:'Mujeres',type:'bar',marker:{color:C.red,line:{width:0}},hovertemplate:'Mujeres: %{y:,.0f}<extra></extra>'}
	            ],{...LO,barmode:'group',bargap:0.2,height:400,xaxis:{...LO.xaxis,tickangle:-45}},CFG);
	        }

	        // Salario por edad y sexo (activos)
	        if(c.salario_edad?.length && document.getElementById('cd-salEdad')){
	            const g=c.salario_edad.map(e=>e.grupo);
	            const cntBy={}; (c.activos_edad||[]).forEach(e=>cntBy[e.grupo]=e);
	            const yH=c.salario_edad.map(e=>Number(cntBy[e.grupo]?.hombres||0)>0?Number(e.hombres||0):null);
	            const yM=c.salario_edad.map(e=>Number(cntBy[e.grupo]?.mujeres||0)>0?Number(e.mujeres||0):null);
	            Plotly.newPlot('cd-salEdad',[
	                {x:g,y:yH,name:'Hombres',type:'scatter',mode:'lines+markers',line:{color:C.navy,width:3,shape:'spline',smoothing:1.15},marker:{size:7},hovertemplate:'Hombres: Gs. %{y:,.0f}<extra></extra>'},
	                {x:g,y:yM,name:'Mujeres',type:'scatter',mode:'lines+markers',line:{color:C.red,width:3,shape:'spline',smoothing:1.15},marker:{size:7},hovertemplate:'Mujeres: Gs. %{y:,.0f}<extra></extra>'}
	            ],{...LO,height:420,xaxis:{...LO.xaxis,tickangle:-45},yaxis:{...LO.yaxis,title:{text:'Guaraníes',font:{size:12}}},margin:{...LO.margin,b:90}},CFG);
	        }

	        // Monto vejez por edad
	        if(c.monto_vejez_edad?.length && document.getElementById('cd-montoVejez')){
	            const g=c.monto_vejez_edad.map(e=>e.grupo);
	            Plotly.newPlot('cd-montoVejez',[
                {x:g,y:c.monto_vejez_edad.map(e=>e.hombres),name:'Hombres',type:'bar',marker:{color:C.navy},hovertemplate:'Hombres: %{y:,.0f} Gs.<extra></extra>'},
                {x:g,y:c.monto_vejez_edad.map(e=>e.mujeres),name:'Mujeres',type:'bar',marker:{color:C.red},hovertemplate:'Mujeres: %{y:,.0f} Gs.<extra></extra>'}
            ],{...LO,barmode:'group',bargap:0.2,height:400,xaxis:{...LO.xaxis,tickangle:-45},yaxis:{...LO.yaxis,title:{text:'Guaraníes',font:{size:12}}}},CFG);
        }
    };
    setTimeout(()=>selCaja(cn[0],document.querySelector('#s-cajas .pill')),150);
}();

	// ======================== SEC: GENERO ========================
	!function(){
	    const s=document.createElement('div'); s.id='s-genero'; s.className='section';
	    s.innerHTML=`
	    <div class="sec-header"><h2>Análisis por Género</h2><p>Distribución de hombres y mujeres en el sistema previsional</p></div>
	    <div class="pill-row">${YRS.map((y,i)=>`<button class="pill${y==='2024'?' on':''}" onclick="selGenYear('${y}',this)">${y}</button>`).join('')}</div>
	    <div class="grid g2">
	        <div class="card"><div class="card-head"><h3>Activos por Sexo — Todas las Cajas</h3></div><div class="card-body" id="g-act"></div></div>
	        <div class="card"><div class="card-head"><h3>Pasivos por Sexo — Todas las Cajas</h3></div><div class="card-body" id="g-pas"></div></div>
	    </div>
	    <div class="grid g2">
	        <div class="card"><div class="card-head"><h3>Participación Femenina en Activos por Caja</h3><div class="tag">% Mujeres</div></div><div class="card-sub">La línea punteada indica paridad de género (50%)</div><div class="card-body" id="g-brecha"></div></div>
	        <div class="card"><div class="card-head"><h3>Participación Femenina en Pasivos por Caja</h3><div class="tag red">% Mujeres</div></div><div class="card-sub">La línea punteada indica paridad de género (50%)</div><div class="card-body" id="g-brecha-pas"></div></div>
	    </div>
	    <div class="grid g1">
	        <div class="card"><div class="card-head"><h3>Brecha Salarial de Género (Activos)</h3><div class="tag">Mujeres/Hombres</div></div><div class="card-sub">Salario promedio ponderado por edad (última tabla disponible)</div><div class="card-body" id="g-sal-gap"></div></div>
	    </div>
	    <div class="grid g1">
	        <div class="card"><div class="card-head"><h3>Evolución de la Participación Femenina en Activos</h3></div><div class="card-sub">Porcentaje de mujeres sobre el total de cotizantes activos por caja</div><div class="card-body" id="g-evo"></div></div>
	    </div>`;
	    M.appendChild(s);

	    setTimeout(()=>{
	        const cn=Object.keys(D).filter(c=>D[c].por_sexo?.activos?.Hombres);

	        const stackL={...LO,barmode:'stack',bargap:0.35,height:400,xaxis:{...LO.xaxis,tickangle:-30},margin:{...LO.margin,b:90}};

	        window.selGenYear=function(year,btn){
	            document.querySelectorAll('#s-genero .pill').forEach(b=>b.classList.remove('on'));
	            if(btn) btn.classList.add('on');

	            const hA=cn.map(c=>Number(D[c].por_sexo?.activos?.Hombres?.[year]||0));
	            const mA=cn.map(c=>Number(D[c].por_sexo?.activos?.Mujeres?.[year]||0));
	            const hP=cn.map(c=>Number(D[c].por_sexo?.pasivos?.Hombres?.[year]||0));
	            const mP=cn.map(c=>Number(D[c].por_sexo?.pasivos?.Mujeres?.[year]||0));

	            Plotly.newPlot('g-act',[
	                {x:cn,y:hA,name:'Hombres',type:'bar',marker:{color:C.navy},hovertemplate:'%{y:,.0f}<extra>Hombres</extra>'},
	                {x:cn,y:mA,name:'Mujeres',type:'bar',marker:{color:C.red},hovertemplate:'%{y:,.0f}<extra>Mujeres</extra>'}
	            ],stackL,CFG);
	            Plotly.newPlot('g-pas',[
	                {x:cn,y:hP,name:'Hombres',type:'bar',marker:{color:C.navy},hovertemplate:'%{y:,.0f}<extra>Hombres</extra>'},
	                {x:cn,y:mP,name:'Mujeres',type:'bar',marker:{color:C.red},hovertemplate:'%{y:,.0f}<extra>Mujeres</extra>'}
	            ],stackL,CFG);

	            // Participación femenina (activos/pasivos)
	            const pctMA=cn.map((c,i)=>{const t=hA[i]+mA[i]; return t>0?(mA[i]/t*100):0;});
	            const pctMP=cn.map((c,i)=>{const t=hP[i]+mP[i]; return t>0?(mP[i]/t*100):0;});
	            const yMaxA=Math.max(60, Math.max(...pctMA)*1.2);
	            const yMaxP=Math.max(60, Math.max(...pctMP)*1.2);
	            const commonBrecha=(pctVals, yMax, color)=>({
	                data:[{
	                    x:cn, y:pctVals, type:'bar',
	                    marker:{color:pctVals.map(p=>p>=50?C.red:C.navy),line:{width:0}},
	                    text:pctVals.map(p=>p.toFixed(1)+'%'), textposition:'outside', textfont:{size:12,weight:700},
	                    hovertemplate:'<b>%{x}</b><br>%{y:.1f}% mujeres<extra></extra>'
	                }],
	                layout:{...LO,height:360,yaxis:{...LO.yaxis,range:[0,yMax],title:{text:'% Mujeres'}},
	                    shapes:[{type:'line',x0:-0.5,x1:cn.length-0.5,y0:50,y1:50,line:{color,width:2,dash:'dash'}}],
	                    annotations:[{x:cn.length-1,y:51.5,text:'<b>Paridad</b>',showarrow:false,font:{size:11,color}}],
	                    margin:{...LO.margin,b:86}}
	            });
	            const ba=commonBrecha(pctMA, yMaxA, C.red);
	            const bp=commonBrecha(pctMP, yMaxP, C.red);
	            Plotly.newPlot('g-brecha',ba.data,ba.layout,CFG);
	            Plotly.newPlot('g-brecha-pas',bp.data,bp.layout,CFG);
	        };

	        // Brecha salarial por caja (activos)
	        const gaps=Object.keys(D).map(c=>{
	            const g=payGapActivos(D[c]);
	            return (g && g.ratio>0)?{c,p:g.ratio*100,avgH:g.avgH,avgM:g.avgM}:null;
	        }).filter(Boolean);
	        if(gaps.length){
	            const yMax=Math.max(110, Math.max(...gaps.map(g=>g.p))*1.2);
	            Plotly.newPlot('g-sal-gap',[{
	                x:gaps.map(g=>g.c), y:gaps.map(g=>g.p), type:'bar',
	                marker:{color:gaps.map(g=>g.p>=100?C.teal:C.red),line:{width:0}},
	                text:gaps.map(g=>g.p.toFixed(1)+'%'), textposition:'outside', textfont:{size:12,weight:700},
	                customdata:gaps.map(g=>[g.avgM,g.avgH]),
	                hovertemplate:'<b>%{x}</b><br>Mujeres: Gs. %{customdata[0]:,.0f}<br>Hombres: Gs. %{customdata[1]:,.0f}<br>Ratio: %{y:.1f}%<extra></extra>'
	            }],{...LO,height:380,yaxis:{...LO.yaxis,title:{text:'Mujeres como % del salario de hombres'},range:[0,yMax]},
	                shapes:[{type:'line',x0:-0.5,x1:gaps.length-0.5,y0:100,y1:100,line:{color:C.gray,width:1.6,dash:'dot'}}],
	                annotations:[{x:gaps[gaps.length-1].c,y:102,text:'<b>100%</b>',showarrow:false,font:{size:10,color:C.gray}}],
	                margin:{...LO.margin,b:86}},CFG);
	        } else {
	            document.getElementById('g-sal-gap').innerHTML='<p style="padding:40px;color:#aaa;text-align:center">Sin datos de salarios por sexo/edad</p>';
	        }

	        // Evo femenina
	        const evoT=cn.map(c=>({
	            x:YRS,
	            y:YRS.map(y=>{
                const h=Number(D[c].por_sexo?.activos?.Hombres?.[y]||0);
                const m=Number(D[c].por_sexo?.activos?.Mujeres?.[y]||0);
                return (h+m)>0?(m/(h+m)*100):0;
            }),
            name:c, mode:'lines+markers', line:{color:CAJA_C[c],width:2.5,shape:'spline',smoothing:1.15}, marker:{size:7}
	        }));
	        Plotly.newPlot('g-evo',evoT,{...LO,height:380,yaxis:{...LO.yaxis,title:{text:'% Mujeres'},range:[0,70]},shapes:[{type:'line',x0:YRS[0],x1:YRS[4],y0:50,y1:50,line:{color:'#adb5bd',width:1.5,dash:'dot'}}]},CFG);

	        // Init
	        const initBtn=document.querySelector('#s-genero .pill-row .pill.on') || document.querySelector('#s-genero .pill-row .pill');
	        if(initBtn) selGenYear('2024', initBtn);
	    },200);
	}();

// ======================== SEC: PIRAMIDES ========================
!function(){
    const s=document.createElement('div'); s.id='s-edades'; s.className='section';
    const cwe=Object.entries(D).filter(([,c])=>c.activos_edad?.length>0);
    s.innerHTML=`
    <div class="sec-header"><h2>Pirámides Poblacionales</h2><p>Estructura etaria de cotizantes activos y jubilados por vejez</p></div>
    <div class="pill-row">${cwe.map(([n],i)=>`<button class="pill${i===0?' on':''}" onclick="selPir('${n}',this)">${n}</button>`).join('')}</div>
	    <div class="grid g2">
	        <div class="card"><div class="card-head"><h3>Pirámide: Cotizantes Activos</h3><div class="tag">Por edad y sexo</div></div><div class="card-sub">Hombres (izquierda) — Mujeres (derecha)</div><div class="card-body" id="p-act"></div></div>
	        <div class="card"><div class="card-head"><h3>Pirámide: Jubilados por Vejez</h3><div class="tag red">Por edad y sexo</div></div><div class="card-sub">Hombres (izquierda) — Mujeres (derecha)</div><div class="card-body" id="p-vej"></div></div>
	    </div>
	    <div class="grid g2">
	        <div class="card"><div class="card-head"><h3>Edad Promedio — Cotizantes Activos</h3><div class="tag">Estimación</div></div><div class="card-sub">Promedio ponderado por grupos de edad</div><div class="card-body" id="p-edadavg"></div></div>
	        <div class="card"><div class="card-head"><h3>Índice de Envejecimiento — Activos 55+</h3><div class="tag red">% del total</div></div><div class="card-sub">Proporción de cotizantes en tramos 55–85+</div><div class="card-body" id="p-55plus"></div></div>
	    </div>`;
    M.appendChild(s);

    window.selPir=function(name,btn){
        document.querySelectorAll('#s-edades .pill').forEach(b=>b.classList.remove('on'));
        if(btn) btn.classList.add('on');
        const c=D[name];

        function drawPyramid(elId, data, colorL, colorR){
            if(!data||!data.length){ document.getElementById(elId).innerHTML='<p style="padding:40px;color:#aaa;text-align:center">Sin datos disponibles</p>'; return; }
            const g=data.map(e=>e.grupo).reverse();
            const hV=data.map(e=>e.hombres).reverse();
            const mV=data.map(e=>e.mujeres).reverse();
            const maxV=Math.max(...hV,...mV)*1.15;

            Plotly.newPlot(elId,[
                {y:g,x:hV.map(v=>-v),name:'Hombres',type:'bar',orientation:'h',marker:{color:colorL,line:{width:0}},
                 customdata:hV,hovertemplate:'<b>%{y}</b><br>Hombres: %{customdata:,.0f}<extra></extra>',
                 text:hV.map(v=>fmt(v)),textposition:'inside',textfont:{size:10,color:'white'},insidetextanchor:'middle'},
                {y:g,x:mV,name:'Mujeres',type:'bar',orientation:'h',marker:{color:colorR,line:{width:0}},
                 hovertemplate:'<b>%{y}</b><br>Mujeres: %{x:,.0f}<extra></extra>',
                 text:mV.map(v=>fmt(v)),textposition:'inside',textfont:{size:10,color:'white'},insidetextanchor:'middle'}
            ],{
                margin:{t:16,b:36,l:72,r:20}, height:Math.max(380, g.length*36),
                paper_bgcolor:'rgba(0,0,0,0)', plot_bgcolor:'rgba(0,0,0,0)',
                font:{family:'DM Sans',size:11,color:'#5a6a7e'},
                barmode:'relative', bargap:0.12,
                xaxis:{range:[-maxV,maxV], showticklabels:false, showgrid:false, zeroline:true, zerolinecolor:'#dee2e6', zerolinewidth:2},
                yaxis:{showgrid:false, tickfont:{size:11,color:'#5a6a7e'}},
                legend:{orientation:'h',y:1.06,x:0.5,xanchor:'center',font:{size:12}}
            },CFG);
        }

	        drawPyramid('p-act', c.activos_edad, C.navy, C.red);
	        drawPyramid('p-vej', c.vejez_edad, C.blue, C.pink);
	    };
	    if(cwe.length>0) setTimeout(()=>selPir(cwe[0][0],document.querySelector('#s-edades .pill')),250);

	    // Indicadores demográficos (por caja)
	    setTimeout(()=>{
	        const cn=cwe.map(([n])=>n);
	        const st=cn.map(n=>({n, s:ageStats(D[n].activos_edad)})).filter(o=>o.s.total);
	        const names=st.map(o=>o.n);
	        const avgs=st.map(o=>o.s.avg);
	        const pct55=st.map(o=>o.s.share55*100);

	        Plotly.newPlot('p-edadavg',[{
	            x:names, y:avgs, type:'bar',
	            marker:{color:names.map(n=>CAJA_C[n]||C.navy),line:{width:0}},
	            text:avgs.map(v=>v.toFixed(1)), textposition:'outside', textfont:{size:12,weight:700},
	            hovertemplate:'<b>%{x}</b><br>Edad promedio: %{y:.1f} años<extra></extra>'
	        }],{...LO,height:360,yaxis:{...LO.yaxis,title:{text:'Años'},rangemode:'tozero'},margin:{...LO.margin,b:86}},CFG);

	        Plotly.newPlot('p-55plus',[{
	            x:names, y:pct55, type:'bar',
	            marker:{color:pct55.map(p=>p>=20?C.red:p>=12?C.amber:C.teal),line:{width:0}},
	            text:pct55.map(v=>v.toFixed(1)+'%'), textposition:'outside', textfont:{size:12,weight:700},
	            hovertemplate:'<b>%{x}</b><br>55+ en activos: %{y:.1f}%<extra></extra>'
	        }],{...LO,height:360,yaxis:{...LO.yaxis,title:{text:'% del total'},range:[0,Math.max(25,Math.max(...pct55)*1.25)]},margin:{...LO.margin,b:86}},CFG);
	    },320);
	}();

// ======================== SEC: SALARIOS ========================
!function(){
    const s=document.createElement('div'); s.id='s-salarios'; s.className='section';
    s.innerHTML=`
    <div class="sec-header"><h2>Salarios y Montos Promedios</h2><p>Remuneraciones de activos y beneficios de jubilados/pensionados</p></div>
    <div class="grid g1">
        <div class="card"><div class="card-head"><h3>Salario Promedio del Cotizante Activo vs Jubilación Promedio</h3><div class="tag">2024 — Guaraníes</div></div><div class="card-body" id="sal-comp"></div></div>
    </div>
    <div class="grid g1">
        <div class="card"><div class="card-head"><h3>Evolución del Salario Promedio del Cotizante Activo</h3><div class="tag">2020–2024</div></div><div class="card-body" id="sal-evo"></div></div>
    </div>
    <div class="grid g1">
        <div class="card"><div class="card-head"><h3>Tasa de Reemplazo: Jubilación Promedio / Salario Activo</h3><div class="tag red">2024</div></div><div class="card-sub">Indica qué porcentaje del salario activo representa la jubilación promedio</div><div class="card-body" id="sal-ratio"></div></div>
    </div>`;
    M.appendChild(s);

    setTimeout(()=>{
        const cws=Object.entries(D).filter(([,c])=>c.salarios?.['2024']);
        const cn=cws.map(([n])=>n);
        const sA=cws.map(([,c])=>Number(c.salarios['2024'].salario_activo||0));
        const sJ=cws.map(([,c])=>Number(c.salarios['2024'].jubilacion_promedio||0));

        // Grouped bar: salario vs jubilacion
        Plotly.newPlot('sal-comp',[
            {x:cn,y:sA,name:'Salario Activo',type:'bar',marker:{color:C.navy},hovertemplate:'%{x}<br>Salario: Gs. %{y:,.0f}<extra></extra>',text:sA.map(v=>'Gs. '+fmtM(v)),textposition:'outside',textfont:{size:10}},
            {x:cn,y:sJ,name:'Jubilación Promedio',type:'bar',marker:{color:C.red},hovertemplate:'%{x}<br>Jubilación: Gs. %{y:,.0f}<extra></extra>',text:sJ.map(v=>'Gs. '+fmtM(v)),textposition:'outside',textfont:{size:10}}
        ],{...LO,barmode:'group',bargap:0.3,height:420,yaxis:{...LO.yaxis,title:{text:'Guaraníes'}},margin:{...LO.margin,b:80,t:36}},CFG);

        // Evolution lines
        const evoT=cws.map(([n,c])=>({
            x:YRS, y:YRS.map(y=>Number(c.salarios?.[y]?.salario_activo||0)),
            name:n, mode:'lines+markers', line:{color:CAJA_C[n],width:2.5,shape:'spline',smoothing:1.15}, marker:{size:7}
        }));
        Plotly.newPlot('sal-evo',evoT,{...LO,height:400,yaxis:{...LO.yaxis,title:{text:'Guaraníes'}}},CFG);

        // Replacement rate
        const ratios=cws.map(([,c])=>{const r=Number(c.salarios['2024'].ratio||0); return r>1?(r*100).toFixed(1):(r*100).toFixed(1);});
        Plotly.newPlot('sal-ratio',[{
            x:cn, y:ratios, type:'bar',
            marker:{color:ratios.map(r=>r>100?C.red:r>80?C.amber:C.teal)},
            text:ratios.map(r=>r+'%'), textposition:'outside', textfont:{size:12,weight:700},
            hovertemplate:'<b>%{x}</b><br>Tasa: %{y:.1f}%<extra></extra>'
        }],{...LO,height:380,yaxis:{...LO.yaxis,title:{text:'% del Salario Activo'},range:[0,Math.max(...ratios.map(Number))*1.2]},shapes:[{type:'line',x0:-0.5,x1:cn.length-0.5,y0:100,y1:100,line:{color:C.red,width:1.5,dash:'dash'}}],annotations:[{x:cn.length-1,y:102,text:'<b>100%</b>',showarrow:false,font:{size:10,color:C.red}}],margin:{...LO.margin,b:80}},CFG);
    },300);
}();

	// ======================== SEC: TABLA COMPARATIVA ========================
	!function(){
	    const s=document.createElement('div'); s.id='s-tabla'; s.className='section';
	    const cn=Object.keys(D);
	    const tA=cn.reduce((s,n)=>s+Number(D[n].activos?.['2024']||0),0);
	    const tP=cn.reduce((s,n)=>s+Number(D[n].pasivos?.['2024']||0),0);
	    let rows='';
	    cn.forEach(n=>{
	        const c=D[n]; const a=Number(c.activos?.['2024']||0); const p=Number(c.pasivos?.['2024']||0);
	        const r=p>0?(a/p).toFixed(2):'—';
	        const sa=c.salarios?.['2024']?.salario_activo?fmtGs(c.salarios['2024'].salario_activo):'—';
	        const sj=c.salarios?.['2024']?.jubilacion_promedio?fmtGs(c.salarios['2024'].jubilacion_promedio):'—';
	        const pA=tA>0?(a/tA*100):0;
	        const pP=tP>0?(p/tP*100):0;
	        rows+=`<tr><td>${n}</td><td>${fmt(a)}</td><td>${pA.toFixed(1)}%</td><td>${fmt(p)}</td><td>${pP.toFixed(1)}%</td><td>${r}</td><td>${sa}</td><td>${sj}</td></tr>`;
	    });
	    rows+=`<tr><td><strong>TOTAL SISTEMA</strong></td><td><strong>${fmt(tA)}</strong></td><td><strong>100.0%</strong></td><td><strong>${fmt(tP)}</strong></td><td><strong>100.0%</strong></td><td><strong>${(tA/tP).toFixed(2)}</strong></td><td>—</td><td>—</td></tr>`;

    s.innerHTML=`
	    <div class="sec-header"><h2>Tabla Comparativa</h2><p>Datos consolidados a diciembre 2024</p></div>
	    <div class="card" style="overflow-x:auto"><div style="padding:24px">
	        <table class="tbl"><thead><tr><th>Caja Previsional</th><th>Activos</th><th>% Activos</th><th>Pasivos</th><th>% Pasivos</th><th>Relación A/P</th><th>Salario Prom. Activo</th><th>Jubilación Prom.</th></tr></thead>
	        <tbody>${rows}</tbody></table>
	    </div></div>
    <div class="grid g1" style="margin-top:24px">
        <div class="card"><div class="card-head"><h3>Ranking de Sostenibilidad: Relación Activos/Pasivos</h3><div class="tag red">2024</div></div><div class="card-sub">Rojo = crítico (&lt;2) · Amarillo = atención (2–4) · Verde = saludable (&gt;4)</div><div class="card-body" id="t-rank"></div></div>
    </div>`;
    M.appendChild(s);

    setTimeout(()=>{
        const sorted=cn.map(n=>{const c=D[n]; const a=Number(c.activos?.['2024']||0); const p=Number(c.pasivos?.['2024']||0); return {n,r:p>0?a/p:0};}).sort((a,b)=>a.r-b.r);
        Plotly.newPlot('t-rank',[{
            y:sorted.map(s=>s.n), x:sorted.map(s=>s.r),
            type:'bar', orientation:'h',
            marker:{color:sorted.map(s=>s.r<2?C.red:s.r<4?C.amber:C.green),line:{width:0}},
            text:sorted.map(s=>s.r.toFixed(2)), textposition:'outside', textfont:{size:13,weight:700},
            hovertemplate:'<b>%{y}</b><br>Ratio: %{x:.2f}<extra></extra>'
        }],{
            margin:{t:16,b:40,l:140,r:60}, height:Math.max(350,sorted.length*50),
            paper_bgcolor:'rgba(0,0,0,0)', plot_bgcolor:'rgba(0,0,0,0)',
            font:{family:'DM Sans',size:12,color:'#5a6a7e'},
            xaxis:{gridcolor:'#eef1f5',title:{text:'Activos por cada Pasivo',font:{size:13}}},
            yaxis:{showgrid:false},
            shapes:[{type:'line',x0:2,x1:2,y0:-0.5,y1:sorted.length-0.5,line:{color:C.red,width:2,dash:'dash'}},{type:'line',x0:4,x1:4,y0:-0.5,y1:sorted.length-0.5,line:{color:C.amber,width:1.5,dash:'dot'}}],
            annotations:[{x:2,y:sorted.length-0.3,text:'<b>Crítico</b>',showarrow:false,yshift:14,font:{size:10,color:C.red}},{x:4,y:sorted.length-0.3,text:'<b>Atención</b>',showarrow:false,yshift:14,font:{size:10,color:C.amber}}]
        },CFG);
    },350);
}();

</script>
</body>
</html>'''

out_path = Path(args.out_path).expanduser() if args.out_path else (Path(__file__).resolve().parent / "index.html")
out_path.parent.mkdir(parents=True, exist_ok=True)
out_path.write_text(html, encoding="utf-8")
print(f"Excel: {excel_path}")
print(f"Dashboard generated: {out_path}")
print(f"Size: {out_path.stat().st_size/1024:.1f} KB")
